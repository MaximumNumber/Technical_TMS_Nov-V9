import csv
import json
import io
import os
from datetime import datetime, time as dtime

from django.conf import settings

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Q, Count
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.utils import timezone

from .models import (
    UnifiedUser, College, Department, CollegeDepartment, DepartmentAcademicPeriod,
    Course, Room, Hall, CollegeRoom, CollegeHall, Professor, LectureSchedule,
    LabSchedule, Notification, ScheduleChangeLog, Role,
)
from .views_analytics import build_analytics_data

DAY_ORDER = ['Saturday', 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday']
DAY_NAMES = {
    'Saturday': 'السبت', 'Sunday': 'الأحد', 'Monday': 'الاثنين',
    'Tuesday': 'الثلاثاء', 'Wednesday': 'الأربعاء', 'Thursday': 'الخميس',
}


def require_college_manager_or_system(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.user_type not in ('college_manager', 'department_head', 'system_manager'):
            messages.error(request, 'ليس لديك صلاحية للوصول لهذه الصفحة')
            return redirect('login')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def get_college_for_user(user):
    if user.user_type == 'system_manager':
        return None
    if user.user_type == 'department_head':
        return user.college
    try:
        role = Role.objects.get(id=user.user_id, role='مدير_كلية')
        return role.college
    except Role.DoesNotExist:
        return user.college


def get_accessible_dept_ids_for_user(user):
    """Return list of accessible dept_ids or None (= all) for system_manager."""
    if user.user_type == 'department_head':
        return [user.department_id] if user.department_id else []
    college = get_college_for_user(user)
    if college:
        return list(CollegeDepartment.objects.filter(college=college).values_list('department_id', flat=True))
    return None


def serialize_lecture(lec):
    return {
        'id': lec.id,
        'course_name': lec.course.course_name,
        'course_code': lec.course.course_code,
        'professor_name': lec.professor.name,
        'department_name': lec.department.name,
        'room_name': lec.room.name,
        'room_code': lec.room.code,
        'day_of_week': lec.day_of_week,
        'day_name': DAY_NAMES.get(lec.day_of_week, lec.day_of_week),
        'start_time': str(lec.start_time),
        'end_time': str(lec.end_time),
        'lecture_type': lec.lecture_type,
        'period_id': lec.period_id,
    }


def serialize_lab(lab):
    return {
        'id': lab.id,
        'course_name': lab.course.course_name,
        'course_code': lab.course.course_code,
        'professor_name': lab.professor.name,
        'department_name': lab.department.name,
        'hall_name': lab.hall.name,
        'hall_code': lab.hall.code,
        'day_of_week': lab.day_of_week,
        'day_name': DAY_NAMES.get(lab.day_of_week, lab.day_of_week),
        'start_time': str(lab.start_time),
        'end_time': str(lab.end_time),
        'group_number': lab.group_number,
        'period_id': lab.period_id,
    }


# ─────────────────────────────────────────────
# 1) CONFLICT DETECTION WITH SUGGESTIONS
# ─────────────────────────────────────────────

def get_conflict_suggestions(department_id, professor_id, room_id, day, start_time, end_time,
                              exclude_id=None, college=None, schedule_type='lecture'):
    """Returns detailed conflict analysis + alternative suggestions."""
    conflicts = []
    suggestions = {}

    base_qs = LectureSchedule.objects.filter(
        day_of_week=day,
        start_time__lt=end_time,
        end_time__gt=start_time
    )
    if exclude_id:
        base_qs = base_qs.exclude(id=exclude_id)

    lab_qs = LabSchedule.objects.filter(
        day_of_week=day,
        start_time__lt=end_time,
        end_time__gt=start_time
    )

    # Check room conflict
    room_conflict = base_qs.filter(room_id=room_id).select_related('course', 'department').first()
    if room_conflict:
        conflicts.append({
            'type': 'room',
            'icon': 'fas fa-door-open',
            'title': 'تعارض في القاعة',
            'message': f'القاعة محجوزة لمادة "{room_conflict.course.course_name}" من قسم {room_conflict.department.name}',
            'detail': f'الوقت: {room_conflict.start_time.strftime("%H:%M")} - {room_conflict.end_time.strftime("%H:%M")}',
        })
        # Suggest available rooms for this time slot
        if college:
            busy_room_ids_lectures = list(base_qs.values_list('room_id', flat=True))
            busy_room_ids_labs = []
            alt_rooms = Room.objects.filter(
                collegeroom__college=college, collegeroom__is_active=True
            ).exclude(id__in=busy_room_ids_lectures).select_related('college')[:5]
        else:
            busy_room_ids = list(base_qs.values_list('room_id', flat=True))
            alt_rooms = Room.objects.exclude(id__in=busy_room_ids)[:5]
        suggestions['alternative_rooms'] = [
            {'id': r.id, 'name': r.name, 'code': r.code, 'capacity': r.capacity}
            for r in alt_rooms
        ]

    # Check professor conflict (lecture + lab)
    prof_lec_conflict = base_qs.filter(professor_id=professor_id).select_related('course', 'department').first()
    prof_lab_conflict = lab_qs.filter(Q(professor_id=professor_id) | Q(assistant_id=professor_id)).select_related('course', 'department').first()
    if prof_lec_conflict or prof_lab_conflict:
        conflict_item = prof_lec_conflict or prof_lab_conflict
        conflicts.append({
            'type': 'professor',
            'icon': 'fas fa-chalkboard-teacher',
            'title': 'تعارض للأستاذ',
            'message': f'الأستاذ مشغول بمادة "{conflict_item.course.course_name}" من قسم {conflict_item.department.name}',
            'detail': f'الوقت: {conflict_item.start_time.strftime("%H:%M")} - {conflict_item.end_time.strftime("%H:%M")}',
        })
        # Suggest available times for this professor on other days
        busy_slots = []
        for lec in LectureSchedule.objects.filter(professor_id=professor_id):
            busy_slots.append((lec.day_of_week, lec.start_time, lec.end_time))
        for lab in LabSchedule.objects.filter(Q(professor_id=professor_id) | Q(assistant_id=professor_id)):
            busy_slots.append((lab.day_of_week, lab.start_time, lab.end_time))

        free_days = []
        for d in DAY_ORDER:
            day_busy = [slot for slot in busy_slots if slot[0] == d]
            if len(day_busy) < 6:
                free_days.append(DAY_NAMES[d])
        suggestions['professor_free_days'] = free_days[:4]

        # Suggest alternative professors from same college
        busy_prof_ids = list(set(
            list(base_qs.values_list('professor_id', flat=True)) +
            list(lab_qs.values_list('professor_id', flat=True))
        ))
        if college:
            alt_profs = Professor.objects.filter(
                Q(college=college) | Q(college_relations__college=college)
            ).exclude(id__in=busy_prof_ids).distinct()[:5]
        else:
            alt_profs = Professor.objects.exclude(id__in=busy_prof_ids)[:5]
        suggestions['alternative_professors'] = [
            {'id': p.id, 'name': p.name, 'position': p.position}
            for p in alt_profs
        ]

    # Check department/student group conflict
    dept_conflict = base_qs.filter(department_id=department_id).select_related('course').first()
    if dept_conflict:
        conflicts.append({
            'type': 'department',
            'icon': 'fas fa-users',
            'title': 'تعارض لمجموعة الطلاب',
            'message': f'للقسم محاضرة أخرى في نفس الوقت: "{dept_conflict.course.course_name}"',
            'detail': f'الوقت: {dept_conflict.start_time.strftime("%H:%M")} - {dept_conflict.end_time.strftime("%H:%M")}',
        })
        # Suggest other time slots on same day (common lecture hours)
        busy_times = [(l.start_time, l.end_time) for l in base_qs.filter(department_id=department_id)]
        common_slots = [
            (dtime(8, 0), dtime(10, 0)),
            (dtime(10, 0), dtime(12, 0)),
            (dtime(12, 0), dtime(14, 0)),
            (dtime(14, 0), dtime(16, 0)),
            (dtime(16, 0), dtime(18, 0)),
        ]
        free_slots = []
        for slot_start, slot_end in common_slots:
            overlap = any(
                slot_start < bt[1] and slot_end > bt[0]
                for bt in busy_times
            )
            if not overlap and not (slot_start < end_time and slot_end > start_time):
                free_slots.append(f'{slot_start.strftime("%H:%M")} - {slot_end.strftime("%H:%M")}')
        suggestions['free_time_slots'] = free_slots[:3]

    return conflicts, suggestions


def get_lab_conflict_suggestions(department_id, professor_id, hall_id, day, start_time, end_time,
                                  exclude_id=None, college=None):
    """Returns detailed conflict analysis + alternative suggestions for lab scheduling."""
    conflicts = []
    suggestions = {}

    base_lec_qs = LectureSchedule.objects.filter(
        day_of_week=day,
        start_time__lt=end_time,
        end_time__gt=start_time,
    )
    base_lab_qs = LabSchedule.objects.filter(
        day_of_week=day,
        start_time__lt=end_time,
        end_time__gt=start_time,
    )
    if exclude_id:
        base_lab_qs = base_lab_qs.exclude(id=exclude_id)

    # Check hall conflict
    hall_conflict = base_lab_qs.filter(hall_id=hall_id).select_related('course', 'department').first()
    if hall_conflict:
        conflicts.append({
            'type': 'hall',
            'icon': 'fas fa-flask',
            'title': 'تعارض في المعمل',
            'message': f'المعمل محجوز لمادة "{hall_conflict.course.course_name}" من قسم {hall_conflict.department.name}',
            'detail': f'الوقت: {hall_conflict.start_time.strftime("%H:%M")} - {hall_conflict.end_time.strftime("%H:%M")}',
        })
        # Suggest alternative halls
        busy_hall_ids = list(base_lab_qs.values_list('hall_id', flat=True))
        if college:
            alt_halls = Hall.objects.filter(
                collegehall__college=college, collegehall__is_active=True
            ).exclude(id__in=busy_hall_ids)[:5]
        else:
            alt_halls = Hall.objects.exclude(id__in=busy_hall_ids)[:5]
        suggestions['alternative_halls'] = [
            {'id': h.id, 'name': h.name, 'code': h.code, 'capacity': h.capacity}
            for h in alt_halls
        ]

    # Check professor conflict (both lectures and labs)
    prof_lec_conflict = base_lec_qs.filter(professor_id=professor_id).select_related('course', 'department').first()
    prof_lab_conflict = base_lab_qs.filter(
        Q(professor_id=professor_id) | Q(assistant_id=professor_id)
    ).select_related('course', 'department').first()

    if prof_lec_conflict or prof_lab_conflict:
        conflict_item = prof_lec_conflict or prof_lab_conflict
        conflicts.append({
            'type': 'professor',
            'icon': 'fas fa-chalkboard-teacher',
            'title': 'تعارض للأستاذ',
            'message': f'الأستاذ مشغول بمادة "{conflict_item.course.course_name}" من قسم {conflict_item.department.name}',
            'detail': f'الوقت: {conflict_item.start_time.strftime("%H:%M")} - {conflict_item.end_time.strftime("%H:%M")}',
        })
        # Suggest free days for professor
        busy_slots = []
        for lec in LectureSchedule.objects.filter(professor_id=professor_id):
            busy_slots.append((lec.day_of_week, lec.start_time, lec.end_time))
        for lab in LabSchedule.objects.filter(Q(professor_id=professor_id) | Q(assistant_id=professor_id)):
            busy_slots.append((lab.day_of_week, lab.start_time, lab.end_time))
        free_days = [DAY_NAMES[d] for d in DAY_ORDER if len([s for s in busy_slots if s[0] == d]) < 6]
        suggestions['professor_free_days'] = free_days[:4]

        # Suggest alternative professors
        busy_prof_ids = list(set(
            list(base_lec_qs.values_list('professor_id', flat=True)) +
            list(base_lab_qs.values_list('professor_id', flat=True))
        ))
        if college:
            alt_profs = Professor.objects.filter(
                Q(college=college) | Q(college_relations__college=college)
            ).exclude(id__in=busy_prof_ids).distinct()[:5]
        else:
            alt_profs = Professor.objects.exclude(id__in=busy_prof_ids)[:5]
        suggestions['alternative_professors'] = [
            {'id': p.id, 'name': p.name, 'position': p.position}
            for p in alt_profs
        ]

    # Check department/student group conflict (existing lab OR lecture)
    dept_lab_conflict = base_lab_qs.filter(department_id=department_id).select_related('course').first()
    dept_lec_conflict = base_lec_qs.filter(department_id=department_id).select_related('course').first()
    if dept_lab_conflict or dept_lec_conflict:
        conflict_item = dept_lab_conflict or dept_lec_conflict
        conflicts.append({
            'type': 'department',
            'icon': 'fas fa-users',
            'title': 'تعارض لمجموعة الطلاب',
            'message': f'للقسم جلسة أخرى في نفس الوقت: "{conflict_item.course.course_name}"',
            'detail': f'الوقت: {conflict_item.start_time.strftime("%H:%M")} - {conflict_item.end_time.strftime("%H:%M")}',
        })
        busy_times = (
            [(l.start_time, l.end_time) for l in base_lab_qs.filter(department_id=department_id)] +
            [(l.start_time, l.end_time) for l in base_lec_qs.filter(department_id=department_id)]
        )
        common_slots = [
            (dtime(8, 0), dtime(10, 0)), (dtime(10, 0), dtime(12, 0)),
            (dtime(12, 0), dtime(14, 0)), (dtime(14, 0), dtime(16, 0)),
            (dtime(16, 0), dtime(18, 0)),
        ]
        free_slots = [
            f'{s.strftime("%H:%M")} - {e.strftime("%H:%M")}'
            for s, e in common_slots
            if not any(s < bt[1] and e > bt[0] for bt in busy_times) and not (s < end_time and e > start_time)
        ]
        suggestions['free_time_slots'] = free_slots[:3]

    return conflicts, suggestions


@login_required
def api_check_lab_conflicts_advanced(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
        dept_id = data.get('department_id')
        prof_id = data.get('professor_id')
        hall_id = data.get('hall_id')
        day = data.get('day_of_week')
        start = data.get('start_time')
        end = data.get('end_time')
        exclude_id = data.get('exclude_id')
        college_id = data.get('college_id')

        start_t = dtime.fromisoformat(start) if start else None
        end_t = dtime.fromisoformat(end) if end else None

        college = None
        if college_id:
            try:
                college = College.objects.get(id=college_id)
            except Exception:
                pass

        conflicts, suggestions = get_lab_conflict_suggestions(
            dept_id, prof_id, hall_id, day, start_t, end_t,
            exclude_id=exclude_id, college=college
        )
        return JsonResponse({
            'conflicts': conflicts,
            'suggestions': suggestions,
            'has_conflicts': len(conflicts) > 0,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_check_conflicts_advanced(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
        dept_id = data.get('department_id')
        prof_id = data.get('professor_id')
        room_id = data.get('room_id')
        day = data.get('day_of_week')
        start = data.get('start_time')
        end = data.get('end_time')
        exclude_id = data.get('exclude_id')
        college_id = data.get('college_id')

        start_t = dtime.fromisoformat(start) if start else None
        end_t = dtime.fromisoformat(end) if end else None

        college = None
        if college_id:
            from .models import College
            try:
                college = College.objects.get(id=college_id)
            except Exception:
                pass

        conflicts, suggestions = get_conflict_suggestions(
            dept_id, prof_id, room_id, day, start_t, end_t,
            exclude_id=exclude_id, college=college
        )
        return JsonResponse({
            'conflicts': conflicts,
            'suggestions': suggestions,
            'has_conflicts': len(conflicts) > 0,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ─────────────────────────────────────────────
# 2) ANALYTICS DASHBOARD
# ─────────────────────────────────────────────

@require_college_manager_or_system
def analytics_dashboard(request):
    user = request.user
    college = get_college_for_user(user)
    colleges = College.objects.all()
    selected_college_id = request.GET.get('college_id')

    if user.user_type == 'system_manager' and selected_college_id:
        try:
            college = College.objects.get(id=selected_college_id)
        except College.DoesNotExist:
            college = None

    data = build_analytics_data(college=college)
    return render(request, 'timetable/analytics_dashboard.html', {
        'analytics': data,
        'college': college,
        'colleges': colleges,
        'selected_college_id': selected_college_id,
        'is_system_manager': user.user_type == 'system_manager',
    })


# ─────────────────────────────────────────────
# 3) NOTIFICATIONS - MARK READ
# ─────────────────────────────────────────────

@login_required
@require_POST
def notifications_mark_read(request):
    notif_id = request.POST.get('notification_id')
    if notif_id:
        Notification.objects.filter(id=notif_id, recipient=request.user).update(is_read=True)
    else:
        Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        unread = Notification.objects.filter(recipient=request.user, is_read=False).count()
        return JsonResponse({'success': True, 'unread_count': unread})
    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def api_unread_count(request):
    count = Notification.objects.filter(recipient=request.user, is_read=False).count()
    return JsonResponse({'unread_count': count})


# ─────────────────────────────────────────────
# 4) CHANGE LOG
# ─────────────────────────────────────────────

def log_schedule_change(action, schedule_obj, changed_by, old_data=None, new_data=None, reason=None):
    schedule_type = 'lecture' if isinstance(schedule_obj, LectureSchedule) else 'lab'
    dept_name = schedule_obj.department.name if schedule_obj.department_id else ''
    course_name = schedule_obj.course.course_name if schedule_obj.course_id else ''
    ScheduleChangeLog.objects.create(
        schedule_type=schedule_type,
        schedule_id=schedule_obj.id,
        action=action,
        changed_by=changed_by,
        old_data=old_data,
        new_data=new_data,
        change_reason=reason,
        department_name=dept_name,
        course_name=course_name,
    )


@require_college_manager_or_system
def changelog_list(request):
    user = request.user
    college = get_college_for_user(user)

    logs_qs = ScheduleChangeLog.objects.select_related('changed_by').order_by('-changed_at')

    if college:
        dept_ids = list(CollegeDepartment.objects.filter(college=college).values_list('department_id', flat=True))
        # Filter by department name
        dept_names = list(Department.objects.filter(id__in=dept_ids).values_list('name', flat=True))
        logs_qs = logs_qs.filter(department_name__in=dept_names)

    # Filters
    action_filter = request.GET.get('action', '')
    type_filter = request.GET.get('type', '')
    search = request.GET.get('search', '')

    if action_filter:
        logs_qs = logs_qs.filter(action=action_filter)
    if type_filter:
        logs_qs = logs_qs.filter(schedule_type=type_filter)
    if search:
        logs_qs = logs_qs.filter(
            Q(course_name__icontains=search) |
            Q(department_name__icontains=search) |
            Q(changed_by__full_name__icontains=search)
        )

    paginator = Paginator(logs_qs, 30)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'timetable/changelog.html', {
        'page_obj': page_obj,
        'logs': page_obj.object_list,
        'college': college,
        'action_filter': action_filter,
        'type_filter': type_filter,
        'search': search,
        'is_system_manager': user.user_type == 'system_manager',
    })


@require_college_manager_or_system
def schedule_restore(request, log_id):
    log_entry = get_object_or_404(ScheduleChangeLog, pk=log_id)

    if request.method == 'POST':
        try:
            if log_entry.action == 'delete' and log_entry.old_data:
                # Restore deleted schedule
                data = log_entry.old_data
                if log_entry.schedule_type == 'lecture':
                    # Check if still deleted
                    if not LectureSchedule.objects.filter(id=log_entry.schedule_id).exists():
                        LectureSchedule.objects.create(
                            id=log_entry.schedule_id,
                            department_id=data.get('department_id'),
                            period_id=data.get('period_id'),
                            course_id=data.get('course_id'),
                            professor_id=data.get('professor_id'),
                            room_id=data.get('room_id'),
                            day_of_week=data.get('day_of_week'),
                            start_time=data.get('start_time'),
                            end_time=data.get('end_time'),
                            lecture_type=data.get('lecture_type', 'lecture'),
                        )
                        messages.success(request, 'تم استعادة المحاضرة المحذوفة بنجاح')
                    else:
                        messages.warning(request, 'المحاضرة موجودة بالفعل')
                elif log_entry.schedule_type == 'lab':
                    if not LabSchedule.objects.filter(id=log_entry.schedule_id).exists():
                        LabSchedule.objects.create(
                            id=log_entry.schedule_id,
                            department_id=data.get('department_id'),
                            period_id=data.get('period_id'),
                            course_id=data.get('course_id'),
                            professor_id=data.get('professor_id'),
                            hall_id=data.get('hall_id'),
                            day_of_week=data.get('day_of_week'),
                            start_time=data.get('start_time'),
                            end_time=data.get('end_time'),
                            group_number=data.get('group_number'),
                        )
                        messages.success(request, 'تم استعادة جدول المعمل المحذوف بنجاح')
                    else:
                        messages.warning(request, 'جدول المعمل موجود بالفعل')

            elif log_entry.action == 'edit' and log_entry.old_data:
                # Revert to old values
                data = log_entry.old_data
                if log_entry.schedule_type == 'lecture':
                    lec = LectureSchedule.objects.filter(id=log_entry.schedule_id).first()
                    if lec:
                        lec.day_of_week = data.get('day_of_week', lec.day_of_week)
                        lec.start_time = data.get('start_time', str(lec.start_time))
                        lec.end_time = data.get('end_time', str(lec.end_time))
                        lec.room_id = data.get('room_id', lec.room_id)
                        lec.professor_id = data.get('professor_id', lec.professor_id)
                        lec.save()
                        messages.success(request, 'تم الرجوع إلى النسخة السابقة بنجاح')
                    else:
                        messages.error(request, 'المحاضرة غير موجودة')
                elif log_entry.schedule_type == 'lab':
                    lab = LabSchedule.objects.filter(id=log_entry.schedule_id).first()
                    if lab:
                        lab.day_of_week = data.get('day_of_week', lab.day_of_week)
                        lab.start_time = data.get('start_time', str(lab.start_time))
                        lab.end_time = data.get('end_time', str(lab.end_time))
                        lab.hall_id = data.get('hall_id', lab.hall_id)
                        lab.professor_id = data.get('professor_id', lab.professor_id)
                        lab.save()
                        messages.success(request, 'تم الرجوع إلى النسخة السابقة بنجاح')
                    else:
                        messages.error(request, 'جدول المعمل غير موجود')
            else:
                messages.warning(request, 'لا يمكن استعادة هذا السجل')

        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء الاستعادة: {str(e)}')

        return redirect('changelog_list')

    return render(request, 'timetable/changelog_restore.html', {'log': log_entry})


# ─────────────────────────────────────────────
# 5) EXPORT / IMPORT
# ─────────────────────────────────────────────

@require_college_manager_or_system
def export_import_page(request):
    user = request.user
    college = get_college_for_user(user)
    return render(request, 'timetable/export_import.html', {
        'college': college,
        'is_system_manager': user.user_type == 'system_manager',
        'day_map': {
            'Saturday': 'السبت', 'Sunday': 'الأحد', 'Monday': 'الاثنين',
            'Tuesday': 'الثلاثاء', 'Wednesday': 'الأربعاء', 'Thursday': 'الخميس',
        },
    })


def public_schedule_export(request):
    """Export public schedule without requiring login."""
    period_id = request.GET.get('period_id')
    fmt = request.GET.get('format', 'csv')

    if not period_id:
        from django.http import HttpResponseBadRequest
        return HttpResponseBadRequest('period_id مطلوب')

    try:
        period = DepartmentAcademicPeriod.objects.select_related('year', 'department').get(pk=period_id)
    except DepartmentAcademicPeriod.DoesNotExist:
        from django.http import HttpResponseNotFound
        return HttpResponseNotFound('الفترة الأكاديمية غير موجودة')

    lectures = LectureSchedule.objects.filter(period_id=period_id).select_related(
        'course', 'professor', 'room', 'department', 'period__year'
    ).order_by('day_of_week', 'start_time')

    labs = LabSchedule.objects.filter(period_id=period_id).select_related(
        'course', 'professor', 'hall', 'department', 'period__year'
    ).order_by('day_of_week', 'start_time')

    period_label = f"{period.department.name}_{period.year.year_name}_فصل{period.semester_type}"

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="schedule_{period_label}.csv"'
        writer = csv.writer(response)
        writer.writerow(['النوع', 'المادة', 'كود المادة', 'الأستاذ', 'القاعة/المعمل', 'اليوم', 'البداية', 'النهاية'])
        for lec in lectures:
            writer.writerow([
                'محاضرة', lec.course.course_name, lec.course.course_code,
                lec.professor.name, lec.room.name,
                DAY_NAMES.get(lec.day_of_week, lec.day_of_week),
                str(lec.start_time)[:5], str(lec.end_time)[:5],
            ])
        for lab in labs:
            writer.writerow([
                'معمل', lab.course.course_name, lab.course.course_code,
                lab.professor.name, lab.hall.name,
                DAY_NAMES.get(lab.day_of_week, lab.day_of_week),
                str(lab.start_time)[:5], str(lab.end_time)[:5],
            ])
        return response

    elif fmt == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'الجدول الدراسي'
        ws.sheet_view.rightToLeft = True

        tms_color = '0A3D47'
        gold_color = 'F5A623'
        header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
        gold_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
        tms_fill = PatternFill('solid', fgColor=tms_color)
        gold_fill = PatternFill('solid', fgColor=gold_color)
        alt_fill = PatternFill('solid', fgColor='EAF4F6')
        alt_fill2 = PatternFill('solid', fgColor='FFF8E7')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append([f'الجدول الدراسي - {period.department.name} / {period.year.year_name} / فصل {period.semester_type}'])
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.font = Font(bold=True, size=14, color=tms_color, name='Arial')
        title_cell.alignment = center
        ws.row_dimensions[1].height = 30
        ws.append([])

        headers = ['النوع', 'المادة', 'كود المادة', 'الأستاذ', 'القاعة/المعمل', 'اليوم', 'البداية', 'النهاية']
        ws.append(headers)
        hdr_row = ws.max_row
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=hdr_row, column=col, value=h)
            cell.font = header_font
            cell.fill = tms_fill
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[hdr_row].height = 22

        for i, lec in enumerate(lectures):
            row = ['محاضرة', lec.course.course_name, lec.course.course_code,
                   lec.professor.name, lec.room.name,
                   DAY_NAMES.get(lec.day_of_week, lec.day_of_week),
                   str(lec.start_time)[:5], str(lec.end_time)[:5]]
            ws.append(row)
            r = ws.max_row
            fill = alt_fill if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                cell.fill = fill
                cell.alignment = center
                cell.border = border

        for i, lab in enumerate(labs):
            row = ['معمل', lab.course.course_name, lab.course.course_code,
                   lab.professor.name, lab.hall.name,
                   DAY_NAMES.get(lab.day_of_week, lab.day_of_week),
                   str(lab.start_time)[:5], str(lab.end_time)[:5]]
            ws.append(row)
            r = ws.max_row
            fill = alt_fill2 if i % 2 == 0 else PatternFill('solid', fgColor='FFFDF5')
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                cell.fill = fill
                cell.alignment = center
                cell.border = border

        col_widths = [12, 28, 14, 22, 18, 14, 10, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="schedule_{period_label}.xlsx"'
        return response

    elif fmt == 'pdf':
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import arabic_reshaper
        from bidi.algorithm import get_display

        _font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Amiri-Regular.ttf')
        _font_bold = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Amiri-Bold.ttf')
        if 'Amiri' not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont('Amiri', _font_path))
        if 'Amiri-Bold' not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont('Amiri-Bold', _font_bold))

        def ar(text):
            if not text:
                return ''
            try:
                return get_display(arabic_reshaper.reshape(str(text)))
            except Exception:
                return str(text)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                rightMargin=1.5*cm, leftMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        tms_color_rl = colors.HexColor('#0A3D47')
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=14,
                                     fontName='Amiri-Bold', textColor=tms_color_rl,
                                     alignment=TA_CENTER, spaceAfter=6)
        sub_style = ParagraphStyle('S', parent=styles['Normal'], fontSize=9,
                                   fontName='Amiri', textColor=colors.grey,
                                   alignment=TA_CENTER, spaceAfter=16)
        elements = [
            Paragraph(ar(f'الجدول الدراسي - {period.department.name}'), title_style),
            Paragraph(ar(f'{period.year.year_name} / فصل {period.semester_type}') +
                      f'   |   {datetime.now().strftime("%Y-%m-%d")}', sub_style),
            Spacer(1, 0.3*cm),
        ]
        col_headers = [ar(h) for h in ['النوع', 'المادة', 'الأستاذ', 'القاعة/المعمل', 'اليوم', 'البداية', 'النهاية']]
        table_data = [col_headers]
        for lec in lectures:
            table_data.append([
                ar('محاضرة'), ar(lec.course.course_name[:28]), ar(lec.professor.name[:20]),
                ar(lec.room.name[:18]), ar(DAY_NAMES.get(lec.day_of_week, lec.day_of_week)),
                str(lec.start_time)[:5], str(lec.end_time)[:5],
            ])
        for lab in labs:
            table_data.append([
                ar('معمل'), ar(lab.course.course_name[:28]), ar(lab.professor.name[:20]),
                ar(lab.hall.name[:18]), ar(DAY_NAMES.get(lab.day_of_week, lab.day_of_week)),
                str(lab.start_time)[:5], str(lab.end_time)[:5],
            ])
        if len(table_data) > 1:
            col_widths_pdf = [2.5*cm, 7*cm, 5*cm, 4.5*cm, 2.8*cm, 2*cm, 2*cm]
            tbl = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), tms_color_rl),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Amiri-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Amiri'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fa')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                ('ROWHEIGHT', (0, 0), (-1, -1), 22),
            ]))
            elements.append(tbl)
        else:
            elements.append(Paragraph(ar('لا توجد جداول مسجلة لهذه الفترة'),
                                      ParagraphStyle('N', parent=styles['Normal'], fontName='Amiri')))
        doc.build(elements)
        buf.seek(0)
        response = HttpResponse(buf.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="schedule_{period_label}.pdf"'
        return response

    from django.http import HttpResponseBadRequest
    return HttpResponseBadRequest('صيغة التصدير غير مدعومة')


@require_college_manager_or_system
def export_lectures_csv(request):
    user = request.user
    college = get_college_for_user(user)

    lectures = LectureSchedule.objects.select_related(
        'course', 'department', 'professor', 'room', 'period', 'period__year'
    ).order_by('department__name', 'day_of_week', 'start_time')

    dept_ids = get_accessible_dept_ids_for_user(user)
    if dept_ids is not None:
        lectures = lectures.filter(department_id__in=dept_ids)

    dept_filter = request.GET.get('dept_id')
    period_filter = request.GET.get('period_id')
    if dept_filter:
        lectures = lectures.filter(department_id=dept_filter)
    if period_filter:
        lectures = lectures.filter(period_id=period_filter)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    filename = f'lectures_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(['القسم', 'السنة الدراسية', 'الفصل', 'المادة', 'كود المادة', 'الأستاذ',
                     'القاعة', 'كود القاعة', 'اليوم', 'وقت البداية', 'وقت النهاية', 'نوع المحاضرة'])

    for lec in lectures:
        writer.writerow([
            lec.department.name,
            lec.period.year.year_name,
            lec.period.semester_type,
            lec.course.course_name,
            lec.course.course_code,
            lec.professor.name,
            lec.room.name,
            lec.room.code,
            DAY_NAMES.get(lec.day_of_week, lec.day_of_week),
            str(lec.start_time),
            str(lec.end_time),
            'محاضرة' if lec.lecture_type == 'lecture' else 'تمرين',
        ])
    return response


@require_college_manager_or_system
def export_labs_csv(request):
    user = request.user
    college = get_college_for_user(user)

    labs = LabSchedule.objects.select_related(
        'course', 'department', 'professor', 'hall', 'period', 'period__year', 'assistant'
    ).order_by('department__name', 'day_of_week', 'start_time')

    dept_ids = get_accessible_dept_ids_for_user(user)
    if dept_ids is not None:
        labs = labs.filter(department_id__in=dept_ids)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    filename = f'labs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(['القسم', 'السنة الدراسية', 'الفصل', 'المادة', 'كود المادة', 'الأستاذ',
                     'المعاون', 'المعمل', 'كود المعمل', 'اليوم', 'وقت البداية', 'وقت النهاية', 'المجموعة'])

    for lab in labs:
        writer.writerow([
            lab.department.name,
            lab.period.year.year_name,
            lab.period.semester_type,
            lab.course.course_name,
            lab.course.course_code,
            lab.professor.name,
            lab.assistant.name if lab.assistant else '',
            lab.hall.name,
            lab.hall.code,
            DAY_NAMES.get(lab.day_of_week, lab.day_of_week),
            str(lab.start_time),
            str(lab.end_time),
            lab.group_number or '',
        ])
    return response


@require_college_manager_or_system
def export_lectures_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    user = request.user
    college = get_college_for_user(user)

    lectures = LectureSchedule.objects.select_related(
        'course', 'department', 'professor', 'room', 'period', 'period__year'
    ).order_by('department__name', 'day_of_week', 'start_time')

    labs = LabSchedule.objects.select_related(
        'course', 'department', 'professor', 'hall', 'period', 'period__year', 'assistant'
    ).order_by('department__name', 'day_of_week', 'start_time')

    dept_ids = get_accessible_dept_ids_for_user(user)
    if dept_ids is not None:
        lectures = lectures.filter(department_id__in=dept_ids)
        labs = labs.filter(department_id__in=dept_ids)

    wb = Workbook()

    # ── Lectures sheet ──
    ws1 = wb.active
    ws1.title = 'جداول المحاضرات'
    ws1.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color='17616c', end_color='17616c', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, name='Arial', size=11)
    alt_fill = PatternFill(start_color='E8F4F6', end_color='E8F4F6', fill_type='solid')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers1 = ['القسم', 'السنة الدراسية', 'الفصل', 'المادة', 'كود المادة',
                'الأستاذ', 'القاعة', 'كود القاعة', 'اليوم', 'وقت البداية', 'وقت النهاية', 'نوع المحاضرة']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws1.row_dimensions[1].height = 25

    for row_idx, lec in enumerate(lectures, 2):
        row_data = [
            lec.department.name, lec.period.year.year_name, f'فصل {lec.period.semester_type}',
            lec.course.course_name, lec.course.course_code, lec.professor.name,
            lec.room.name, lec.room.code, DAY_NAMES.get(lec.day_of_week, lec.day_of_week),
            str(lec.start_time)[:5], str(lec.end_time)[:5],
            'محاضرة' if lec.lecture_type == 'lecture' else 'تمرين',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.alignment = center
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    col_widths1 = [20, 15, 10, 25, 12, 20, 15, 12, 12, 12, 12, 12]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Labs sheet ──
    ws2 = wb.create_sheet(title='جداول المعامل')
    ws2.sheet_view.rightToLeft = True

    lab_header_fill = PatternFill(start_color='f0a500', end_color='f0a500', fill_type='solid')
    alt_fill2 = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid')

    headers2 = ['القسم', 'السنة الدراسية', 'الفصل', 'المادة', 'كود المادة',
                'الأستاذ', 'المعاون', 'المعمل', 'كود المعمل', 'اليوم', 'وقت البداية', 'وقت النهاية', 'المجموعة']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = lab_header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws2.row_dimensions[1].height = 25

    for row_idx, lab in enumerate(labs, 2):
        row_data = [
            lab.department.name, lab.period.year.year_name, f'فصل {lab.period.semester_type}',
            lab.course.course_name, lab.course.course_code, lab.professor.name,
            lab.assistant.name if lab.assistant else '',
            lab.hall.name, lab.hall.code, DAY_NAMES.get(lab.day_of_week, lab.day_of_week),
            str(lab.start_time)[:5], str(lab.end_time)[:5], lab.group_number or '',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.alignment = center
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill2

    col_widths2 = [20, 15, 10, 25, 12, 20, 20, 15, 12, 12, 12, 12, 10]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'schedules_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@require_college_manager_or_system
def export_schedule_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER
    import arabic_reshaper
    from bidi.algorithm import get_display

    _font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Amiri-Regular.ttf')
    _font_bold = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Amiri-Bold.ttf')
    if 'Amiri' not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont('Amiri', _font_path))
    if 'Amiri-Bold' not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont('Amiri-Bold', _font_bold))

    def ar(text):
        if not text:
            return ''
        try:
            return get_display(arabic_reshaper.reshape(str(text)))
        except Exception:
            return str(text)

    user = request.user
    college = get_college_for_user(user)

    lectures = LectureSchedule.objects.select_related(
        'course', 'department', 'professor', 'room', 'period', 'period__year'
    ).order_by('department__name', 'day_of_week', 'start_time')

    dept_ids = get_accessible_dept_ids_for_user(user)
    if dept_ids is not None:
        lectures = lectures.filter(department_id__in=dept_ids)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    tms_color = colors.HexColor('#17616c')
    alt_bg = colors.HexColor('#f5f7fa')

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Title'], fontSize=16,
                                 fontName='Amiri-Bold', textColor=tms_color,
                                 alignment=TA_CENTER, spaceAfter=10)
    sub_style = ParagraphStyle('SubStyle', parent=styles['Normal'], fontSize=10,
                               fontName='Amiri', textColor=colors.grey,
                               alignment=TA_CENTER, spaceAfter=20)

    elements = []

    college_name = college.name if college else 'جميع الكليات'
    elements.append(Paragraph(ar(f'جدول المحاضرات - {college_name}'), title_style))
    elements.append(Paragraph(f'{datetime.now().strftime("%Y-%m-%d %H:%M")}', sub_style))
    elements.append(Spacer(1, 0.3*cm))

    col_headers = [ar(h) for h in ['القسم', 'المادة', 'الأستاذ', 'القاعة', 'اليوم', 'البداية', 'النهاية', 'النوع']]
    table_data = [col_headers]

    for lec in lectures:
        table_data.append([
            ar(lec.department.name[:20]),
            ar(lec.course.course_name[:25]),
            ar(lec.professor.name[:20]),
            lec.room.code,
            ar(DAY_NAMES.get(lec.day_of_week, lec.day_of_week)),
            str(lec.start_time)[:5],
            str(lec.end_time)[:5],
            ar('محاضرة' if lec.lecture_type == 'lecture' else 'تمرين'),
        ])

    if len(table_data) > 1:
        col_widths = [5*cm, 6*cm, 5*cm, 2.5*cm, 2.5*cm, 2*cm, 2*cm, 2.5*cm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), tms_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Amiri-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Amiri'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, alt_bg]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('ROWHEIGHT', (0, 0), (-1, -1), 22),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph(ar('لا توجد محاضرات مسجلة'),
                                  ParagraphStyle('N', parent=styles['Normal'], fontName='Amiri')))

    doc.build(elements)
    buffer.seek(0)

    filename = f'schedule_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@require_college_manager_or_system
def import_schedule_template(request):
    """Download blank CSV template for importing schedules."""
    schedule_type = request.GET.get('type', 'lectures')
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="import_template_{schedule_type}.csv"'
    writer = csv.writer(response)
    if schedule_type == 'labs':
        writer.writerow(['department_id', 'period_id', 'course_id', 'professor_id',
                         'hall_id', 'day_of_week', 'start_time', 'end_time', 'group_number', 'assistant_id'])
        writer.writerow(['1', '1', '1', '1', '1', 'Saturday', '08:00', '10:00', '1', ''])
    else:
        writer.writerow(['department_id', 'period_id', 'course_id', 'professor_id',
                         'room_id', 'day_of_week', 'start_time', 'end_time', 'lecture_type'])
        writer.writerow(['1', '1', '1', '1', '1', 'Saturday', '08:00', '10:00', 'lecture'])
    return response


@require_college_manager_or_system
def import_schedule_csv(request):
    if request.method != 'POST':
        return redirect('export_import_page')

    user = request.user
    college = get_college_for_user(user)
    schedule_type = request.POST.get('schedule_type', 'lectures')
    csv_file = request.FILES.get('csv_file')

    if not csv_file:
        messages.error(request, 'يرجى اختيار ملف CSV')
        return redirect('export_import_page')

    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'يرجى رفع ملف بصيغة CSV فقط')
        return redirect('export_import_page')

    try:
        content = csv_file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)

        success_count = 0
        error_rows = []

        for i, row in enumerate(rows, 2):
            try:
                if schedule_type == 'labs':
                    group_num = row.get('group_number', '').strip()
                    assistant_id = row.get('assistant_id', '').strip()
                    lab = LabSchedule.objects.create(
                        department_id=int(row['department_id']),
                        period_id=int(row['period_id']),
                        course_id=int(row['course_id']),
                        professor_id=int(row['professor_id']),
                        hall_id=int(row['hall_id']),
                        day_of_week=row['day_of_week'].strip(),
                        start_time=row['start_time'].strip(),
                        end_time=row['end_time'].strip(),
                        group_number=int(group_num) if group_num else None,
                        assistant_id=int(assistant_id) if assistant_id else None,
                    )
                    log_schedule_change('add', lab, user, new_data=serialize_lab(lab))
                else:
                    lec = LectureSchedule.objects.create(
                        department_id=int(row['department_id']),
                        period_id=int(row['period_id']),
                        course_id=int(row['course_id']),
                        professor_id=int(row['professor_id']),
                        room_id=int(row['room_id']),
                        day_of_week=row['day_of_week'].strip(),
                        start_time=row['start_time'].strip(),
                        end_time=row['end_time'].strip(),
                        lecture_type=row.get('lecture_type', 'lecture').strip(),
                    )
                    log_schedule_change('add', lec, user, new_data=serialize_lecture(lec))
                success_count += 1
            except Exception as e:
                error_rows.append(f'صف {i}: {str(e)}')

        if success_count:
            messages.success(request, f'تم استيراد {success_count} سجل بنجاح')
        if error_rows:
            for err in error_rows[:5]:
                messages.warning(request, err)
            if len(error_rows) > 5:
                messages.warning(request, f'...و {len(error_rows) - 5} أخطاء إضافية')

    except Exception as e:
        messages.error(request, f'خطأ في قراءة الملف: {str(e)}')

    return redirect('export_import_page')
